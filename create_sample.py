"""Создание тестового Excel-файла спецификации (9 колонок)."""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


def create_sample():
    wb = Workbook()
    ws = wb.active
    ws.title = "Спецификация"

    headers = [
        "Поз.",
        "Наименование и техническая характеристика",
        "Тип, марка, обозначение документа, опросного листа",
        "Код продукции",
        "Поставщик",
        "Ед. измерения",
        "Кол.",
        "Масса 1 ед., кг",
        "Примечание",
    ]

    header_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.border = thin_border

    data = [
        ["1", "Кондиционер настенный сплит-система, холодопроизводительность 3.5 кВт",
         "Daikin FTXB35C", "48 5130", "Daikin Industries", "шт.", "2", "32.5", ""],
        ["2", "Вентилятор канальный круглый, d=160мм, расход воздуха 420 м3/ч",
         "Systemair K 160 M", "48 6100", "Systemair", "шт.", "3", "5.8", "С регулятором скорости"],
        ["3", "Воздуховод круглый из оцинкованной стали, d=160мм",
         "ВК-160, ГОСТ 24751-81", "57 3300", "Вентпром", "м.п.", "36", "2.1", ""],
        ["4", "Решётка вентиляционная приточная, 300×150мм, алюминиевая",
         "АМН 300×150", "48 6812", "Арктос", "шт.", "12", "0.8", "RAL 9016"],
        ["5", "Клапан обратный круглый, d=160мм",
         "RSK 160", "48 6210", "Systemair", "шт.", "3", "0.6", ""],
        ["6", "Фильтр воздушный кассетный, 490×490×48мм, класс G4",
         "ФВК-490-490-48-G4", "48 6430", "Фильтр-М", "шт.", "6", "1.2", "Замена каждые 3 мес."],
        ["7", "Шумоглушитель круглый, d=160мм, L=600мм",
         "CSA 160/600", "48 6500", "Systemair", "шт.", "3", "3.4", ""],
        ["8", "Гибкая вставка круглая, d=160мм, L=150мм",
         "DEC 160/150", "48 6510", "Systemair", "шт.", "6", "0.3", ""],
        ["9", "Теплоизоляция для воздуховодов, толщина 25мм, самоклеящаяся",
         "K-Flex ST 25мм", "57 4000", "K-Flex", "м2", "24", "0.5", ""],
        ["10", "Диффузор потолочный круглый, d=200мм",
         "ДПУ-200", "48 6814", "Арктос", "шт.", "9", "0.5", "Регулируемый"],
        ["11", "Электронагреватель канальный круглый, d=160мм, 2.4 кВт",
         "CB 160-2.4 230V/1", "48 6310", "Systemair", "шт.", "1", "2.8", ""],
        ["12", "Автоматика управления приточной установкой",
         "Контроллер Breezart CP-JL", "42 1700", "Breezart", "компл.", "1", "1.5", "С пультом ДУ"],
    ]

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    widths = [6, 45, 30, 14, 18, 10, 8, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 40

    output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "resources", "sample_input.xlsx")
    wb.save(output)
    print(f"Created: {output}")


if __name__ == "__main__":
    create_sample()
