"""
Импорт данных спецификации из Excel (.xlsx).
Автодетект заголовков, определение типа строк, округление чисел.
"""

import re
from pathlib import Path
from openpyxl import load_workbook
from core.data_model import SpecificationRow, RowType


class ExcelReaderError(Exception):
    pass

class FileFormatError(ExcelReaderError):
    pass

class HeaderNotFoundError(ExcelReaderError):
    pass


_HEADER_PATTERNS = [
    ("наименование",    "name"),
    ("тип",             "type_brand"),
    ("марка",           "type_brand"),
    ("обозначение док", "type_brand"),
    ("код продукции",   "product_code"),
    ("код",             "product_code"),
    ("окп",             "product_code"),
    ("поставщик",       "supplier"),
    ("завод",           "supplier"),
    ("изготовит",       "supplier"),
    ("масса",           "mass_unit_kg"),
    ("вес",             "mass_unit_kg"),
    ("ед. изм",         "unit"),
    ("ед. ",            "unit"),
    ("кол",             "quantity"),
    ("количество",      "quantity"),
    ("примечан",        "notes"),
    ("поз",             "position"),
]

# Только "name" обязательна — "position" может отсутствовать (автонумерация)
_REQUIRED_FIELDS = {"name"}

# Числовые поля, в которых нужно округлять
_NUMERIC_FIELDS = {"quantity", "mass_unit_kg"}


def _match_header(text: str) -> str | None:
    if not text:
        return None
    text_lower = text.strip().lower()
    for pattern, field_name in _HEADER_PATTERNS:
        if pattern in text_lower:
            return field_name
    return None


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _round_number(text: str) -> str:
    """Округляет длинные десятичные числа до 2 знаков после запятой."""
    if not text:
        return text
    # Проверяем: число с длинной дробной частью (>3 знаков)
    try:
        val = float(text.replace(",", "."))
        # Если это целое число — не трогаем
        if val == int(val) and "." not in text and "," not in text:
            return text
        # Округляем до 2 знаков
        rounded = round(val, 2)
        # Убираем лишние нули: 3.10 → 3.1, но 3.00 → 3
        if rounded == int(rounded):
            return str(int(rounded))
        return str(rounded)
    except (ValueError, OverflowError):
        return text


def _detect_row_type(row_data: dict[str, str], prev_type: RowType | None) -> RowType:
    """
    Определяет тип строки:
    - CATEGORY: 1-2 заполненных текстовых поля (без числовых данных в других)
                и это первая такая строка (нет предшествующей категории)
    - SUBCATEGORY: аналогично, но идёт после CATEGORY
    - DATA: 3+ заполненных полей или есть числовые значения
    """
    # Считаем заполненные поля (непустые)
    text_fields = ["position", "name", "type_brand", "product_code", "supplier"]
    num_fields = ["quantity", "mass_unit_kg"]

    filled_text = sum(1 for f in text_fields if row_data.get(f, "").strip())
    filled_num = sum(1 for f in num_fields if row_data.get(f, "").strip())
    filled_unit = 1 if row_data.get("unit", "").strip() and row_data.get("unit", "") != "шт." else 0
    filled_notes = 1 if row_data.get("notes", "").strip() else 0

    total_filled = filled_text + filled_num + filled_unit + filled_notes

    # Если 3+ заполненных или есть числовые значения с данными — это DATA
    if total_filled >= 3 or filled_num >= 1:
        return RowType.DATA

    # 1-2 заполненных текстовых поля (только position и/или name)
    if filled_text <= 2 and filled_num == 0:
        # Если до этого не было категории — это CATEGORY
        if prev_type is None or prev_type == RowType.DATA:
            return RowType.CATEGORY
        # Если перед нами была CATEGORY — это SUBCATEGORY
        elif prev_type == RowType.CATEGORY:
            return RowType.SUBCATEGORY
        # После SUBCATEGORY — ещё одна SUBCATEGORY
        elif prev_type == RowType.SUBCATEGORY:
            return RowType.SUBCATEGORY

    return RowType.DATA


def read_excel(file_path: str | Path) -> tuple[list[SpecificationRow], list[str]]:
    """Читает Excel, определяет типы строк, округляет числа."""
    path = Path(file_path)
    if not path.exists():
        raise FileFormatError(f"Файл не найден: {path}")
    if path.suffix.lower() not in (".xlsx", ".xls"):
        raise FileFormatError(f"Неподдерживаемый формат: {path.suffix}")

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        raise FileFormatError(f"Не удалось открыть файл: {e}")

    try:
        ws = wb.active
        warnings: list[str] = []

        # Поиск заголовков
        col_mapping: dict[int, str] = {}
        header_row_idx = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=False), start=1):
            mapping = {}
            matched_fields = set()
            for cell in row:
                text = _cell_to_str(cell.value)
                field_name = _match_header(text)
                if field_name and field_name not in matched_fields:
                    mapping[cell.column - 1] = field_name
                    matched_fields.add(field_name)

            if _REQUIRED_FIELDS.issubset(matched_fields):
                col_mapping = mapping
                header_row_idx = row_idx
                break

        if header_row_idx is None:
            raise HeaderNotFoundError(
                "Не найдена строка заголовков. Ожидаются колонки: "
                "Поз., Наименование, Тип/марка, Код, Поставщик, Ед.изм., "
                "Кол., Масса, Примечание"
            )

        rows: list[SpecificationRow] = []
        all_rows = list(ws.iter_rows(min_row=header_row_idx + 1, values_only=True))
        prev_type: RowType | None = None

        for row_values in all_rows:
            if not row_values or all(v is None for v in row_values):
                continue

            row_data = {}
            for col_idx, field_name in col_mapping.items():
                if col_idx < len(row_values):
                    val = _cell_to_str(row_values[col_idx])
                    # Округляем числа в числовых полях
                    if field_name in _NUMERIC_FIELDS:
                        val = _round_number(val)
                    row_data[field_name] = val

            # Пропускаем строки совсем без данных
            if not any(row_data.get(f, "").strip() for f in
                       ["position", "name", "type_brand", "product_code", "supplier",
                        "quantity", "mass_unit_kg", "notes"]):
                continue

            # Определяем тип строки
            row_type = _detect_row_type(row_data, prev_type)
            prev_type = row_type

            spec_row = SpecificationRow(
                position=row_data.get("position", ""),
                name=row_data.get("name", ""),
                type_brand=row_data.get("type_brand", ""),
                product_code=row_data.get("product_code", ""),
                supplier=row_data.get("supplier", ""),
                unit=row_data.get("unit", ""),
                quantity=row_data.get("quantity", ""),
                mass_unit_kg=row_data.get("mass_unit_kg", ""),
                notes=row_data.get("notes", ""),
                row_type=row_type,
            )
            rows.append(spec_row)

        if not rows:
            warnings.append("Файл не содержит строк данных после заголовков.")

        found_fields = set(col_mapping.values())
        optional_missing = {"type_brand", "product_code", "supplier", "unit",
                            "quantity", "mass_unit_kg", "notes"} - found_fields
        if optional_missing:
            warnings.append(f"Не найдены колонки: {', '.join(optional_missing)}")

        return rows, warnings
    finally:
        wb.close()
